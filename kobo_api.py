import os
import sys
import traceback
import unicodedata
from collections import defaultdict
from datetime import datetime
from urllib import response

import pandas as pd
import requests
from dotenv import load_dotenv
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(BASE_DIR, ".env")

load_dotenv(ENV_PATH)
def crear_sesion():
    token = os.getenv("APIKEY")

    if not token:
        raise ValueError("No se encontró la variable de entorno APIKEY.")

    session = requests.Session()
    session.headers.update({
        "Authorization": f"Token {token}"
    })

    return session

""" obtener el uid del formualriod e donde se obtendra la información de los sumideros """
def obtener_uid(session):
    url = "https://kf.kobotoolbox.org/api/v2/assets/"

    response = session.get(url)
    response.raise_for_status()

    datos = response.json()

    return datos["results"][0]["uid"]

""" def obtener_formulario(session, uid):
    url = f"https://kf.kobotoolbox.org/api/v2/assets/{uid}/data/"

    resultados = []

    while url:
        response = session.get(url)
        response.raise_for_status()

        data = response.json()
        resultados.extend(data["results"])

        # URL de la siguiente página (None cuando ya no hay más)
        url = data["next"]
    return resultados


 """

def obtener_formulario(session, uid): 
    url = f"https://kf.kobotoolbox.org/api/v2/assets/{uid}/data/?limit=200" 
    response = session.get(url) 
    response.raise_for_status() 
    return response.json()
  
def procesar_datos(formulario):
    sumideros=[]
    sumideros_sin_id = []
    #for sumidero in formulario:
    for sumidero in formulario["results"]:
        foto_antes = ""
        foto_despues = ""
        # Buscar las URLs de las fotos del antes y despues
        for adj in sumidero.get("_attachments", []):
            if adj.get("question_xpath") == "TOME_UNA_FOTO_DEL_SU_ES_DEL_MANTENIMIENTO":
                foto_antes = adj.get("download_url", "")

            elif adj.get("question_xpath") == "TOME_UNA_FOTO_DEL_SU_ES_DEL_MANTENIMIENTO_001":
                foto_despues = adj.get("download_url", "")  
        registro = {
            "ID_ELEMENTO": (
                f"{(sumidero.get('ID') if sumidero.get('ID') is not None else sumidero.get('ID_sumidero'))}S"
                if (sumidero.get('ID') if sumidero.get('ID') is not None else sumidero.get('ID_sumidero')) is not None
                else ""
            ),
            "SUBTIPO": sumidero.get("SUBTIPO").upper().replace("SIF_N", "SIFÓN") if sumidero.get("SUBTIPO") else "",
            "BARRIO": normalizar_barrio(sumidero.get("BARRIO")),            
            "DIRECCIÓN": sumidero.get("DIRECCI_N").upper() if sumidero.get("DIRECCI_N") else "",
            "ESTADO": sumidero.get("ESTADO").upper() if sumidero.get("ESTADO") else "",
            "NUMERO REJAS": float(sumidero.get("NUMERO_DE_REJAS", "0").replace(",", ".")) if sumidero.get("NUMERO_DE_REJAS") else "",
            "PROFUNDIDAD": float(sumidero.get("PROFUNDIDAD", "0").replace(",", ".")),
            "MATERIAL TUBO": sumidero.get("MATERIAL").replace("_", " ").upper() if sumidero.get("MATERIAL") else "",
            "DIAMETRO TUBO": float(sumidero.get("DIAMETRO_TUBO", "0").replace(",", ".")) if sumidero.get("DIAMETRO_TUBO") else "",
            "ENTREGA": sumidero.get("ENTREGA").upper() if sumidero.get("ENTREGA") else "",
            "OBSERVACIONES": sumidero.get("OBSERVACIONES", "").upper() if sumidero.get("OBSERVACIONES") else "",
            "FECHA LIMPIEZA": sumidero.get("today").upper() if sumidero.get("today") else "",
            "OPERARIO": sumidero.get("OPERARIO").upper() if sumidero.get("OPERARIO") else "",
            "URL_FOTOS_ANTES": foto_antes,
            "URL_FOTOS_DESPUÉS": foto_despues
        }
        if sumidero.get("ID", ""):
            sumideros.append(registro)
        else:
            sumideros_sin_id.append(registro)
        
    resultado= sumideros + sumideros_sin_id

    return resultado

def normalizar_barrio(barrio):
    if not barrio:
        return ""

    barrio = str(barrio).strip().upper()

    return "".join(
        c for c in unicodedata.normalize("NFD", barrio)
        if unicodedata.category(c) != "Mn"
    )

def crear_excel(resultado_formulario, session, carpetaSeleccionada, crearArchivo):
    df = pd.DataFrame(resultado_formulario)

    columnas = [
        "ID_ELEMENTO",
        "SUBTIPO",
        "BARRIO",
        "DIRECCIÓN",
        "ESTADO",
        "NUMERO REJAS",
        "PROFUNDIDAD",
        "MATERIAL TUBO",
        "DIAMETRO TUBO",
        "ENTREGA",
        "OBSERVACIONES",
        "FECHA LIMPIEZA",
        "OPERARIO",
    ]

    df = df[columnas]
    año = datetime.now().year
    mes = datetime.now().month
    periodo = 1 if mes <= 6 else 2

    carpeta_semestre = f"Mantenimiento_Sumideros_{año}_{periodo}"
    listaBarrios=[]
    # --------- DIFERENCIA ENTRE CREAR Y ACTUALIZAR ---------
    if carpeta_semestre[0:23] != os.path.basename(carpetaSeleccionada)[0:23] and crearArchivo:
        rutaCarpeta = os.path.join(carpetaSeleccionada, carpeta_semestre)
        carpeta_fotos = os.path.join(rutaCarpeta, "Fotos")
        archivo_excel = os.path.join(rutaCarpeta, "Mantenimiento.xlsx")
        os.makedirs(carpeta_fotos, exist_ok=True)

        with pd.ExcelWriter(
            archivo_excel,
            engine="openpyxl"
        ) as writer:

            for barrio, grupo in df.groupby("BARRIO"):
                nombre_hoja = str(barrio)[:31]
                listaBarrios.append(nombre_hoja)
                grupo.to_excel(
                    writer,
                    sheet_name=nombre_hoja,
                    index=False
                )
        formatear_hoja(archivo_excel, listaBarrios)
        formatear_excel(archivo_excel)
        descargar_fotos(resultado_formulario,session,carpeta_fotos)
        insertar_imagenes(archivo_excel,resultado_formulario,carpeta_fotos)

    elif not crearArchivo:
        archivo_excel = os.path.join(carpetaSeleccionada, "Mantenimiento.xlsx")
        carpeta_fotos = os.path.join(carpetaSeleccionada, "Fotos")
        wb = load_workbook(archivo_excel)

        for barrio, grupo in df.groupby("BARRIO"): 
            nombre_hoja = str(barrio)[:31]
            # Si la hoja no existe, crearla
            if nombre_hoja not in wb.sheetnames:
                ws = wb.create_sheet(nombre_hoja)
                ws.append(list(df.columns))
                ids_existentes = set()
                listaBarrios.append(nombre_hoja)
            else:
                ws = wb[nombre_hoja]
                ids_existentes = set()
                # Desde la fila 7 porque tu formato comienza allí
                for fila in ws.iter_rows(min_row=7, values_only=True):

                    id_elemento = fila[0]
                    if id_elemento is not None:
                        ids_existentes.add(str(id_elemento))

            
            # Buscar la siguiente fila libre
            fila_destino = ws.max_row + 1

            for _, registro in grupo.iterrows():

                id_actual = str(registro["ID_ELEMENTO"])

                if id_actual not in ids_existentes:
                    for columna, valor in enumerate(registro.tolist(), start=1):
                        ws.cell(row=fila_destino, column=columna).value = valor

                    fila_destino += 1
        wb.save(archivo_excel)
        formatear_hoja(archivo_excel,listaBarrios)
        formatear_excel(archivo_excel)
        descargar_fotos(resultado_formulario,session,carpeta_fotos)
        insertar_imagenes(archivo_excel,resultado_formulario,carpeta_fotos)

def descargar_fotos(resultado_formulario, session, carpeta_fotos):
    for registro in resultado_formulario:

        barrio = normalizar_barrio(registro["BARRIO"])
        id_elemento = registro.get("ID_ELEMENTO")

        if  id_elemento == "" or id_elemento is None:
            id_elemento = registro.get("DIRECCIÓN")

        carpeta = os.path.join(carpeta_fotos, barrio)

        os.makedirs(carpeta, exist_ok=True)

        fotos = {
            "antes": registro["URL_FOTOS_ANTES"],
            "despues": registro["URL_FOTOS_DESPUÉS"]
        }

        for tipo, url in fotos.items():

            if url == "":
                continue

            nombre = f"{id_elemento}_{tipo}.jpg"

            ruta = os.path.join(carpeta, nombre)

            if not os.path.exists(ruta):
                try:
                    r = session.get(url)
                    if r.status_code == 200:
                        with open(ruta, "wb") as f:
                            f.write(r.content)
                    else:
                        print(f"Error al descargar {nombre}: código {r.status_code}")
                except Exception as e:
                    print(f"Error al procesar la URL {url}: {e}")
   
def formatear_hoja(archivo_excel, listaBarrios):
    wb = load_workbook(archivo_excel)
    for ws in wb.worksheets:
        if listaBarrios and ws.title in listaBarrios:
        
            # Insertar espacio superior para encabezado
            ws.insert_rows(1, amount=5)
            ws.insert_cols(14, amount=2)
            ws.cell(row=6, column=14).value = "Foto antes"
            ws.cell(row=6, column=15).value = "Foto después"

            # Título principal
            ws.merge_cells("F2:J2")
            ws["F2"] = "FORMATO LIMPIEZA DE SUMIDEROS - MANTENIMIENTO"
            ws["F2"].font = Font(
                bold=True,
                size=12,
                color="1F1F1F"
            )
            ws["F2"].alignment = Alignment(
                horizontal="center"
            )

            # Contrato
            ws.merge_cells("F3:J3")
            ws["F3"] = "CONTRATO: 2026"
            ws["F3"].font = Font(
                bold=True
            )

            # Contratista
            ws.merge_cells("F4:J4")
            ws["F4"] = "CONTRATISTA: CONSTRUCTORA GAF SAS"
            ws["F4"].font = Font(
                bold=True
            )
            # Encabezado tabla
            relleno = PatternFill(
                "solid",
                fgColor="C6E0B4"
            )

            fuente = Font(
                bold=True,
                color="006100"
            )
            for cell in ws[6]:
                cell.fill = relleno
                cell.font = fuente
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",  
                ) 
    wb.save(archivo_excel)


def formatear_excel(archivo_excel):

    wb = load_workbook(archivo_excel)

    for ws in wb.worksheets:
        for columna in ["F", "G", "I"]:
            for celda in ws[columna][6:]: 
                if isinstance(celda.value, (int, float)):
                    celda.number_format = "0.00"
 
        borde = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        # Aplicar desde A6 hasta O(última fila)
        for fila in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=15):
            ws.row_dimensions[fila[0].row].height = 110
            for celda in fila:
                celda.border = borde
                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
        # Altura encabezado
        ws.row_dimensions[6].height = 45
        ws.row_dimensions[6].width = 55
    
        # Estado
        for fila in range(7, ws.max_row+1):

            estado = ws[f"E{fila}"]

            if estado.value:

                valor = str(estado.value).upper()


                if valor in ["IRREGULAR","REGULAR"]:

                    estado.fill = PatternFill(
                        "solid",
                        fgColor="FFF2CC"
                    )
                    estado.font = Font( color="C76E00")

                elif valor == "MALO":

                    estado.fill = PatternFill(
                        "solid",
                        fgColor="F4CCCC"
                    )
                    estado.font = Font(
                        color="FF0000"     # Letra roja
                    )


                estado.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
        ws.auto_filter.ref = f"A6:M6"

    wb.save(archivo_excel)

def insertar_imagenes(archivo_excel, resultado_formulario, carpeta_fotos):

    wb = load_workbook(archivo_excel)

    barrios = agrupar_por_barrio(resultado_formulario)

    for nombre_hoja, registros in barrios.items():

        ws = wb[str(nombre_hoja)[:31]]

        col_antes, col_despues = obtener_columnas_imagenes(ws)
        

        for fila, registro in enumerate(registros, start=7):
            if not existe_imagen(ws, col_antes, col_despues, fila):
                insertar_imagen_fila(
                    ws,
                    fila,
                    registro,
                    carpeta_fotos,
                    col_antes,
                    col_despues
                )

        ajustar_anchos_columnas(ws, col_antes, col_despues)

    wb.save(archivo_excel)

def agrupar_por_barrio(resultado_formulario):

    barrios = defaultdict(list)

    for registro in resultado_formulario:
        barrios[
            normalizar_barrio(registro["BARRIO"])
        ].append(registro)

    return barrios

def obtener_columnas_imagenes(ws):

    col_antes = ws.max_column - 1
    col_despues = ws.max_column

    return col_antes, col_despues

def obtener_rutas_imagenes(registro, carpeta_fotos):

    barrio = registro["BARRIO"]

    id_elemento = (
        registro["ID_ELEMENTO"]
        if registro["ID_ELEMENTO"]
        else registro["DIRECCIÓN"]
    )

    ruta_antes = os.path.join(
        carpeta_fotos,
        barrio,
        f"{id_elemento}_antes.jpg"
    )

    ruta_despues = os.path.join(
        carpeta_fotos,
        barrio,
        f"{id_elemento}_despues.jpg"
    )

    return ruta_antes, ruta_despues

def insertar_imagen_fila(
    ws,
    fila,
    registro,
    carpeta_fotos,
    col_antes,
    col_despues
):
    ruta_antes, ruta_despues = obtener_rutas_imagenes(
        registro,
        carpeta_fotos
    )

    if os.path.exists(ruta_antes):

        insertar_imagen(
            ws,
            ruta_antes,
            fila,
            col_antes
        )

    if os.path.exists(ruta_despues):

        insertar_imagen(
            ws,
            ruta_despues,
            fila,
            col_despues
        )

def insertar_imagen(ws, ruta, fila, columna):

    img = Image(ruta)

    img.width = 80
    img.height = 140

    centrar_imagen(ws, img, fila, columna)

    ws.add_image(img)

def centrar_imagen(ws, img, fila, columna):

    ancho_columna_px = 140
    alto_fila_px = 120

    offset_x = max((ancho_columna_px - img.width) // 2, 0)
    offset_y = max((alto_fila_px - img.height) // 2, 0)

    marker = AnchorMarker(
        col=columna - 1,
        row=fila - 1,
        colOff=pixels_to_EMU(offset_x),
        rowOff=pixels_to_EMU(offset_y)
    )

    img.anchor = OneCellAnchor(
    _from=marker,
    ext=XDRPositiveSize2D(
        pixels_to_EMU(img.width),
        pixels_to_EMU(img.height)
    )
)

def ajustar_anchos_columnas(
    ws,
    col_antes,
    col_despues
):

    for col in ws.columns:

        letra = get_column_letter(col[0].column)

        if letra in (
            get_column_letter(col_antes),
            get_column_letter(col_despues)
        ):

            ws.column_dimensions[letra].width = 20

        else:

            longitud = max(
                len(str(c.value)) if c.value else 0
                for c in col
            )

            ws.column_dimensions[letra].width = min(
                longitud + 3,
                25
            )

def existe_imagen(ws, col_antes, col_despues, fila):
    """
    Verifica si ya existe una imagen en la fila indicada,
    ya sea en la columna ANTES o en la columna DESPUÉS.
    """
    fila_idx = fila - 1
    for img in ws._images:
        if img.anchor._from.row == fila_idx and (
            img.anchor._from.col == col_antes or
            img.anchor._from.col == col_despues
        ):
            return True

    return False

def proceso(rutaCarpeta, crearArchivo):
    try:
        session = crear_sesion()
        uid = obtener_uid(session)

        datos = obtener_formulario(session, uid)
        resultado = procesar_datos(datos)
        crear_excel(resultado, session, rutaCarpeta, crearArchivo)

        return True

    except Exception as e:
        print(f"Error en proceso: {e}")
        return False
